"""Tests for core Excel generation logic."""

import pytest
from pathlib import Path
from openpyxl import load_workbook

from core import audit_delivery_readiness, build_issue_report, ExcelBuilderError


def test_build_uxeval_report(temp_output_dir, mock_issues):
    """Test building a uxeval template report."""
    output_path = temp_output_dir / "uxeval_report.xlsx"

    result = build_issue_report(
        issues=mock_issues,
        output_path=str(output_path),
        template="uxeval",
    )

    assert result["issue_report"]["sheet_count"] == 3
    assert Path(result["issue_report"]["path"]).exists()
    assert Path(result["html_report"]["path"]).exists()
    assert Path(result["evidence_pack"]["path"]).is_dir()
    assert (Path(result["evidence_pack"]["path"]) / "manifest.json").exists()

    # Verify workbook can be loaded
    wb = load_workbook(output_path)
    assert len(wb.sheetnames) == 3
    assert "问题清单" in wb.sheetnames
    assert "摘要" in wb.sheetnames
    assert "原则覆盖" in wb.sheetnames

    # Verify 问题清单 sheet content
    ws_issues = wb["问题清单"]
    assert ws_issues.cell(1, 1).value == "ID"
    assert ws_issues.cell(2, 1).value == "I-001"
    assert ws_issues.cell(2, 2).value == "登录按钮位置不明显"
    assert ws_issues.cell(2, 3).value == "critical"

    # Verify 摘要 sheet content
    ws_summary = wb["摘要"]
    assert ws_summary.cell(1, 1).value == "严重等级"
    assert ws_summary.cell(2, 1).value == "critical"
    assert ws_summary.cell(2, 2).value == 1  # 1 critical issue

    wb.close()


def test_build_design_acceptance_report(temp_output_dir, mock_issues):
    """Test building a design-acceptance template report."""
    output_path = temp_output_dir / "design_acceptance_report.xlsx"

    result = build_issue_report(
        issues=mock_issues,
        output_path=str(output_path),
        template="design-acceptance",
    )

    assert result["issue_report"]["sheet_count"] == 3
    assert Path(result["issue_report"]["path"]).exists()
    assert Path(result["html_report"]["path"]).exists()
    assert Path(result["evidence_pack"]["path"]).is_dir()

    # Verify workbook can be loaded
    wb = load_workbook(output_path)
    assert len(wb.sheetnames) == 3
    assert "差异清单" in wb.sheetnames
    assert "页面汇总" in wb.sheetnames
    assert "组件汇总" in wb.sheetnames

    wb.close()


def test_build_competitor_report(temp_output_dir, mock_issues):
    """Test building a competitor template report."""
    output_path = temp_output_dir / "competitor_report.xlsx"

    result = build_issue_report(
        issues=mock_issues,
        output_path=str(output_path),
        template="competitor",
    )

    assert result["issue_report"]["sheet_count"] == 2
    assert Path(result["issue_report"]["path"]).exists()
    assert Path(result["html_report"]["path"]).exists()
    assert Path(result["evidence_pack"]["path"]).is_dir()

    # Verify workbook can be loaded
    wb = load_workbook(output_path)
    assert len(wb.sheetnames) == 2
    assert "功能对比矩阵" in wb.sheetnames
    assert "维度评分" in wb.sheetnames

    wb.close()


def test_unknown_template_raises_error(temp_output_dir, mock_issues):
    """Test that unknown template raises ExcelBuilderError."""
    output_path = temp_output_dir / "report.xlsx"

    with pytest.raises(ExcelBuilderError, match="Unknown template"):
        build_issue_report(
            issues=mock_issues,
            output_path=str(output_path),
            template="unknown-template",
        )


def test_invalid_output_path_raises_error(mock_issues):
    """Test that invalid output path raises ExcelBuilderError."""
    output_path = "/nonexistent/directory/report.xlsx"

    with pytest.raises(ExcelBuilderError, match="Output directory does not exist"):
        build_issue_report(
            issues=mock_issues,
            output_path=output_path,
            template="uxeval",
        )


def test_empty_issues_for_uxeval_raises_error(temp_output_dir):
    """Test that empty issues list raises error for uxeval template."""
    output_path = temp_output_dir / "report.xlsx"

    with pytest.raises(ExcelBuilderError, match="Issues list cannot be empty"):
        build_issue_report(
            issues=[],
            output_path=str(output_path),
            template="uxeval",
        )


def test_file_overwrite(temp_output_dir, mock_issues):
    """Test that existing file is overwritten."""
    output_path = temp_output_dir / "report.xlsx"

    # Create first report
    result1 = build_issue_report(
        issues=mock_issues,
        output_path=str(output_path),
        template="uxeval",
    )
    assert Path(result1["issue_report"]["path"]).exists()

    # Overwrite with second report
    result2 = build_issue_report(
        issues=mock_issues[:2],  # Fewer issues
        output_path=str(output_path),
        template="uxeval",
    )
    assert Path(result2["issue_report"]["path"]).exists()

    # Verify the file was overwritten
    wb = load_workbook(output_path)
    ws = wb["问题清单"]
    # Should have 2 issues + 1 header = 3 rows
    assert ws.max_row == 3

    wb.close()


def test_severity_color_coding(temp_output_dir, mock_issues):
    """Test that severity levels have correct color coding."""
    output_path = temp_output_dir / "report.xlsx"

    build_issue_report(
        issues=mock_issues,
        output_path=str(output_path),
        template="uxeval",
    )

    wb = load_workbook(output_path)
    ws = wb["问题清单"]

    # Check critical issue (row 2)
    critical_cell = ws.cell(2, 3)
    assert critical_cell.fill.start_color.rgb.endswith("FFCCCC")

    # Check major issue (row 3)
    major_cell = ws.cell(3, 3)
    assert major_cell.fill.start_color.rgb.endswith("FFE5CC")

    # Check minor issue (row 4)
    minor_cell = ws.cell(4, 3)
    assert minor_cell.fill.start_color.rgb.endswith("FFFFCC")

    # Check suggestion (row 5)
    suggestion_cell = ws.cell(5, 3)
    assert suggestion_cell.fill.start_color.rgb.endswith("E6E6E6")

    wb.close()


def test_principle_aggregation(temp_output_dir, mock_issues):
    """Test that principles are correctly aggregated."""
    output_path = temp_output_dir / "report.xlsx"

    build_issue_report(
        issues=mock_issues,
        output_path=str(output_path),
        template="uxeval",
    )

    wb = load_workbook(output_path)
    ws = wb["原则覆盖"]

    # H1 appears in 3 issues (I-001, I-003, I-005)
    # Find H1 row
    h1_count = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == "H1":
            h1_count = row[1]
            break

    assert h1_count == 3

    wb.close()


def test_build_uxeval_report_returns_stage_output_contract(temp_output_dir, mock_issues):
    """Stage 7 tool contract exposes same-named structured outputs."""
    result = build_issue_report(
        issues=mock_issues,
        output_dir=str(temp_output_dir),
        template="uxeval",
    )

    assert set(result) == {"issue_report", "html_report", "evidence_pack"}
    assert result["issue_report"]["id"] == "issue_report"
    assert result["issue_report"]["type"] == "issue_report"
    assert result["html_report"]["id"] == "html_report"
    assert result["html_report"]["type"] == "html_report"
    assert result["evidence_pack"]["id"] == "evidence_pack"
    assert result["evidence_pack"]["type"] == "evidence_pack"


def test_delivery_audit_rejects_llm_claimed_final_when_main_issue_is_not_verified(temp_output_dir):
    issues = [
        {
            "id": "I-001",
            "title": "Primary CTA is buried",
            "severity": "major",
            "description": "CTA is below the fold.",
            "user_impact": "User misses the primary action.",
            "suggestion": "Move CTA above the fold.",
            "evidence_refs": ["E-001"],
            "evidence_basis": ["ocr CTA cue"],
            "verification_status": "verified",
        },
        {
            "id": "I-002",
            "title": "Error state is inconsistent",
            "severity": "major",
            "description": "State text differs across pages.",
            "user_impact": "Users cannot trust result feedback.",
            "suggestion": "Unify state feedback copy.",
            "evidence_refs": ["E-002"],
            "evidence_basis": ["markdown description"],
            "verification_status": "needs_verification",
        },
    ]
    result = audit_delivery_readiness(
        issues=issues,
        unverified_issues=[],
        evidence_assessment={
            "delivery_status": "final_delivery_ready",
            "required_actions": [],
            "missing_coverage": [],
            "verification_gaps": [],
            "coverage_summary": {},
        },
        delivery_assessment={
            "delivery_status": "final_delivery_ready",
        },
        output_dir=str(temp_output_dir),
    )

    audited = result["audited_delivery_assessment"]
    assert audited["delivery_status"] == "fallback_safe"
    assert audited["final_delivery_ready"] is False
    assert audited["delivery_readiness_breakdown"]["fallback_gate_pass"] is True
    assert "issue_qualification" in audited["delivery_readiness_breakdown"]["failing_final_gates"]
    assert "failed deterministic verification" in " ".join(audited["audit_failures"])
    benchmark = result["client_mode_benchmark_summary"]
    assert benchmark["delivery_status"] == "fallback_safe"
    assert benchmark["metrics"]["trust_metrics"]["unverified_leakage_rate"] > 0.0
    bundle = Path(result["delivery_audit_bundle"]["path"])
    assert (bundle / "bounded_issue_pass.md").exists()
    assert (bundle / "unverified_issues.json").exists()
    assert (bundle / "supplement_request.md").exists()
    assert (temp_output_dir / "benchmark" / "client_mode_benchmark_summary.json").exists()
    assert (temp_output_dir / "benchmark" / "client_mode_benchmark_summary.md").exists()


def test_delivery_audit_creates_bounded_fallback_package(temp_output_dir):
    issues = [
        {
            "id": "I-001",
            "title": "Primary CTA is buried",
            "severity": "major",
            "description": "CTA is below the fold.",
            "user_impact": "User misses the primary action.",
            "suggestion": "Move CTA above the fold.",
            "evidence_refs": ["E-001"],
            "evidence_basis": ["ocr CTA cue"],
            "verification_status": "verified",
        }
    ]
    result = audit_delivery_readiness(
        issues=issues,
        unverified_issues=[
            {
                "id": "I-099",
                "title": "Export success feedback may be missing",
                "severity": "major",
                "blocked_by": ["缺导出成功状态截图"],
            }
        ],
        evidence_assessment={
            "delivery_status": "fallback_safe",
            "required_actions": ["补导出成功状态截图"],
            "missing_coverage": ["关键状态覆盖不足"],
            "verification_gaps": ["导出成功状态未覆盖"],
            "fusion_summary": {
                "trusted_page_mappings": [{"relative_path": "screen-01.png"}],
                "provisional_mappings": [{"relative_path": "screen-02.png"}],
                "conflicting_evidence_groups": [{"relative_path": "screen-03.png"}],
                "unresolved_ambiguities": [{"relative_path": "screen-04.png"}],
            },
            "coverage_summary": {
                "capture_mission_version": "2026-05-25",
                "missing_state_categories": ["success"],
                "missing_tasks": ["导出成功"],
                "fusion_conflicting_paths": ["screen-03.png"],
                "fusion_unresolved_paths": ["screen-04.png"],
            },
        },
        delivery_assessment={
            "delivery_status": "fallback_safe",
        },
        capture_mission={
            "mission_version": "2026-05-25",
            "critical_flows": ["导出报表"],
            "capture_order": ["导出入口", "导出成功反馈"],
            "final_delivery_pass_line": ["must_capture_states 覆盖率 >= 80%。"],
            "fallback_pass_line": ["must_capture_states 覆盖率 >= 40%。"],
        },
        output_dir=str(temp_output_dir),
    )

    audited = result["audited_delivery_assessment"]
    assert audited["delivery_status"] == "fallback_safe"
    assert audited["capture_mission_version"] == "2026-05-25"
    assert audited["fusion_summary"]["conflict_count"] == 1
    assert audited["delivery_readiness_breakdown"]["fallback_gate_pass"] is True
    assert "trusted_evidence_sufficiency" in audited["delivery_readiness_breakdown"]["failing_final_gates"]
    benchmark = result["client_mode_benchmark_summary"]
    assert benchmark["delivery_status"] == "fallback_safe"
    assert benchmark["metrics"]["success_metrics"]["fallback_safe_rate"] == 1.0
    bundle = Path(result["delivery_audit_bundle"]["path"])
    supplement_request = (bundle / "supplement_request.md").read_text(encoding="utf-8")
    bounded_issue_pass = (bundle / "bounded_issue_pass.md").read_text(encoding="utf-8")
    assert "Delivery readiness breakdown" in supplement_request
    assert "success" in supplement_request
    assert "Capture Mission pass lines" in supplement_request
    assert "Primary CTA is buried" in bounded_issue_pass


def test_delivery_audit_allows_final_ready_when_all_rules_pass(temp_output_dir):
    issues = [
        {
            "id": "I-001",
            "title": "Primary CTA is buried",
            "severity": "major",
            "description": "CTA is below the fold.",
            "user_impact": "User misses the primary action.",
            "suggestion": "Move CTA above the fold.",
            "evidence_refs": ["E-001"],
            "evidence_basis": ["ocr CTA cue", "markdown description"],
            "verification_status": "verified",
        }
    ]
    result = audit_delivery_readiness(
        issues=issues,
        unverified_issues=[],
        evidence_assessment={
            "delivery_status": "final_delivery_ready",
            "required_actions": [],
            "missing_coverage": [],
            "verification_gaps": [],
            "coverage_summary": {},
        },
        delivery_assessment={"delivery_status": "final_delivery_ready"},
        output_dir=str(temp_output_dir),
    )

    audited = result["audited_delivery_assessment"]
    assert audited["delivery_status"] == "final_delivery_ready"
    assert audited["final_delivery_ready"] is True
    assert audited["delivery_readiness_breakdown"]["final_gate_pass"] is True
    assert audited["delivery_readiness_breakdown"]["failing_final_gates"] == []
    benchmark = result["client_mode_benchmark_summary"]
    assert benchmark["delivery_status"] == "final_delivery_ready"
    assert benchmark["metrics"]["success_metrics"]["final_delivery_ready_rate"] == 1.0
    assert benchmark["metrics"]["trust_metrics"]["unverified_leakage_rate"] == 0.0
    bundle = Path(result["delivery_audit_bundle"]["path"])
    assert (bundle / "audited_delivery_assessment.json").exists()


def test_delivery_audit_downgrades_claimed_final_when_critical_path_final_gate_fails(temp_output_dir):
    issues = [
        {
            "id": "I-001",
            "title": "Primary CTA is buried",
            "severity": "major",
            "description": "CTA is below the fold.",
            "user_impact": "User misses the primary action.",
            "suggestion": "Move CTA above the fold.",
            "evidence_refs": ["E-001"],
            "evidence_basis": ["ocr CTA cue"],
            "verification_status": "verified",
        }
    ]
    result = audit_delivery_readiness(
        issues=issues,
        unverified_issues=[],
        evidence_assessment={
            "delivery_status": "final_delivery_ready",
            "required_actions": ["补设置页关键状态截图"],
            "missing_coverage": ["P1 设置页主链路未过 final line"],
            "verification_gaps": ["设置页主链路缺少高置信覆盖"],
            "coverage_summary": {
                "final_delivery_missing_critical_paths": ["[P1] 设置页"],
                "fallback_missing_critical_paths": [],
                "missing_critical_pages": ["设置页"],
            },
            "critical_path_coverage_summary": {
                "critical_paths": [
                    {
                        "path_name": "设置页",
                        "priority": "P1",
                        "final_delivery_pass": False,
                        "fallback_pass": True,
                    }
                ],
                "failing_final_paths": ["[P1] 设置页"],
                "failing_fallback_paths": [],
            },
        },
        delivery_assessment={"delivery_status": "final_delivery_ready"},
        capture_mission={
            "mission_version": "2026-05-25",
            "critical_flows": ["设置保存"],
            "final_delivery_pass_line": ["所有 P0/P1 critical path 必须过线。"],
            "fallback_pass_line": ["所有 P0 critical path 至少要过 fallback 线。"],
        },
        output_dir=str(temp_output_dir),
    )

    audited = result["audited_delivery_assessment"]
    assert audited["delivery_status"] == "fallback_safe"
    assert audited["final_delivery_ready"] is False
    assert audited["critical_path_coverage_summary"]["failing_final_paths"] == ["[P1] 设置页"]
    assert "critical_path_coverage" in audited["delivery_readiness_breakdown"]["failing_final_gates"]
    bundle = Path(result["delivery_audit_bundle"]["path"])
    supplement_request = (bundle / "supplement_request.md").read_text(encoding="utf-8")
    assert "[P1] 设置页" in supplement_request


def test_delivery_audit_downgrades_claimed_final_when_trusted_evidence_gate_fails(temp_output_dir):
    issues = [
        {
            "id": "I-001",
            "title": "Primary CTA is buried",
            "severity": "major",
            "description": "CTA is below the fold.",
            "user_impact": "User misses the primary action.",
            "suggestion": "Move CTA above the fold.",
            "evidence_refs": ["E-001"],
            "evidence_basis": ["ocr CTA cue"],
            "verification_status": "verified",
        }
    ]
    result = audit_delivery_readiness(
        issues=issues,
        unverified_issues=[],
        evidence_assessment={
            "delivery_status": "final_delivery_ready",
            "required_actions": ["补设置页成功态截图"],
            "missing_coverage": [],
            "verification_gaps": ["trusted evidence still insufficient for 设置页成功态"],
            "coverage_summary": {
                "final_delivery_page_coverage_ratio": 0.75,
                "final_delivery_state_coverage_ratio": 0.5,
                "final_delivery_trusted_mapping_count": 2,
                "final_delivery_missing_planned_states": ["设置页:success"],
                "fallback_missing_critical_paths": [],
                "final_delivery_missing_critical_paths": [],
            },
        },
        delivery_assessment={"delivery_status": "final_delivery_ready"},
        output_dir=str(temp_output_dir),
    )

    audited = result["audited_delivery_assessment"]
    assert audited["delivery_status"] == "fallback_safe"
    assert audited["final_delivery_ready"] is False
    assert "trusted_evidence_sufficiency" in audited["delivery_readiness_breakdown"]["failing_final_gates"]


def test_delivery_audit_clarification_residue_blocks_final_release(temp_output_dir):
    issues = [
        {
            "id": "I-001",
            "title": "Primary CTA is buried",
            "severity": "major",
            "description": "CTA is below the fold.",
            "user_impact": "User misses the primary action.",
            "suggestion": "Move CTA above the fold.",
            "evidence_refs": ["E-001"],
            "evidence_basis": ["ocr CTA cue"],
            "verification_status": "verified",
        }
    ]
    result = audit_delivery_readiness(
        issues=issues,
        unverified_issues=[],
        evidence_assessment={
            "delivery_status": "final_delivery_ready",
            "required_actions": ["先确认 1 张歧义截图的页面/状态"],
            "missing_coverage": [],
            "verification_gaps": ["clarification residue still remains"],
            "coverage_summary": {
                "final_delivery_page_coverage_ratio": 1.0,
                "final_delivery_state_coverage_ratio": 1.0,
                "final_delivery_trusted_mapping_count": 5,
                "clarification_needed_count": 1,
                "clarification_unlocks_final_count": 1,
                "clarification_relative_paths": ["IMG2303.png"],
                "fusion_unresolved_ambiguity_count": 1,
                "fallback_missing_critical_paths": [],
                "final_delivery_missing_critical_paths": [],
            },
        },
        delivery_assessment={"delivery_status": "final_delivery_ready"},
        output_dir=str(temp_output_dir),
    )

    audited = result["audited_delivery_assessment"]
    assert audited["delivery_status"] == "fallback_safe"
    assert audited["final_delivery_ready"] is False
    assert "clarification_residue" in audited["delivery_readiness_breakdown"]["failing_final_gates"]


__all__ = []
