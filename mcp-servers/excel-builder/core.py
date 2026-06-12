"""Core Excel report generation logic for DesignOS.

Pure functions that build Excel workbooks from Issue data structures.
All functions are type-annotated and side-effect-free (except file I/O).
"""

from __future__ import annotations

import html
import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class ExcelBuilderError(Exception):
    """Base exception for excel-builder errors."""

    pass


# Severity level color mapping
SEVERITY_COLORS = {
    "critical": "FFCCCC",  # Red
    "major": "FFE5CC",  # Orange
    "minor": "FFFFCC",  # Yellow
    "suggestion": "E6E6E6",  # Gray
}

HEADER_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
HEADER_FONT = Font(bold=True)
_BENCHMARK_DIRNAME = "benchmark"
_BENCHMARK_JSON_FILENAME = "client_mode_benchmark_summary.json"
_BENCHMARK_MD_FILENAME = "client_mode_benchmark_summary.md"
_BENCHMARK_CONTRACT_VERSION = "2026-05-27"


def _require_worksheet(sheet: object, *, label: str) -> Worksheet:
    """Return a concrete worksheet or raise a typed builder error."""
    if not isinstance(sheet, Worksheet):
        raise ExcelBuilderError(f"{label} is not a writable worksheet")
    return cast(Worksheet, sheet)


def _auto_adjust_column_width(ws: Worksheet, max_width: int = 50) -> None:
    """Auto-adjust column widths based on content length."""
    for column in ws.columns:
        max_length = 0
        column_index = column[0].column
        if isinstance(column_index, str):
            column_letter = column_index
        elif isinstance(column_index, int):
            column_letter = get_column_letter(column_index)
        else:
            continue

        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def _apply_header_style(ws: Worksheet, row_num: int = 1) -> None:
    """Apply header styling to the specified row."""
    for cell in ws[row_num]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_uxeval_template(wb: Workbook, issues: list[dict[str, Any]]) -> None:
    """Build the uxeval template with 3 sheets."""
    # Sheet 1: 问题清单
    ws_issues = _require_worksheet(wb.active, label="uxeval issues sheet")
    ws_issues.title = "问题清单"

    headers = ["ID", "标题", "严重等级", "原则", "描述", "证据", "建议"]
    ws_issues.append(headers)
    _apply_header_style(ws_issues)

    for issue in issues:
        severity = issue.get("severity", "")
        principle_ids = ", ".join(issue.get("principle_ids", []))
        evidence_refs = ", ".join(issue.get("evidence_refs", []))

        row = [
            issue.get("id", ""),
            issue.get("title", ""),
            severity,
            principle_ids,
            issue.get("description", ""),
            evidence_refs,
            issue.get("suggestion", ""),
        ]
        ws_issues.append(row)

        # Apply severity color to the severity column
        row_num = ws_issues.max_row
        severity_cell = ws_issues.cell(row=row_num, column=3)
        if severity in SEVERITY_COLORS:
            severity_cell.fill = PatternFill(
                start_color=SEVERITY_COLORS[severity],
                end_color=SEVERITY_COLORS[severity],
                fill_type="solid",
            )

    _auto_adjust_column_width(ws_issues)

    # Sheet 2: 摘要
    ws_summary = _require_worksheet(wb.create_sheet("摘要"), label="uxeval summary sheet")
    ws_summary.append(["严重等级", "数量"])
    _apply_header_style(ws_summary)

    severity_counts: dict[str, int] = {}
    for issue in issues:
        severity = issue.get("severity", "unknown")
        severity_counts[severity] = severity_counts.get(severity, 0) + 1

    for severity in ["critical", "major", "minor", "suggestion"]:
        count = severity_counts.get(severity, 0)
        ws_summary.append([severity, count])

    _auto_adjust_column_width(ws_summary)

    # Sheet 3: 原则覆盖
    ws_principles = _require_worksheet(wb.create_sheet("原则覆盖"), label="uxeval principles sheet")
    ws_principles.append(["原则ID", "问题数量"])
    _apply_header_style(ws_principles)

    principle_counts: dict[str, int] = {}
    for issue in issues:
        for principle_id in issue.get("principle_ids", []):
            principle_counts[principle_id] = principle_counts.get(principle_id, 0) + 1

    for principle_id, count in sorted(principle_counts.items()):
        ws_principles.append([principle_id, count])

    _auto_adjust_column_width(ws_principles)


def _build_design_acceptance_template(wb: Workbook, issues: list[dict[str, Any]]) -> None:
    """Build the design-acceptance template with 3 sheets."""
    # Sheet 1: 差异清单
    ws_diff = _require_worksheet(wb.active, label="design acceptance diff sheet")
    ws_diff.title = "差异清单"

    headers = ["ID", "页面", "差异类型", "设计值", "实现值", "偏差量"]
    ws_diff.append(headers)
    _apply_header_style(ws_diff)

    for issue in issues:
        row = [
            issue.get("id", ""),
            issue.get("module_id", ""),
            issue.get("title", ""),
            "",  # 设计值 placeholder
            "",  # 实现值 placeholder
            "",  # 偏差量 placeholder
        ]
        ws_diff.append(row)

    _auto_adjust_column_width(ws_diff)

    # Sheet 2: 页面汇总
    ws_page = _require_worksheet(wb.create_sheet("页面汇总"), label="design acceptance page sheet")
    ws_page.append(["页面", "问题数量"])
    _apply_header_style(ws_page)

    page_counts: dict[str, int] = {}
    for issue in issues:
        page = issue.get("module_id", "unknown")
        page_counts[page] = page_counts.get(page, 0) + 1

    for page, count in sorted(page_counts.items()):
        ws_page.append([page, count])

    _auto_adjust_column_width(ws_page)

    # Sheet 3: 组件汇总
    ws_component = _require_worksheet(
        wb.create_sheet("组件汇总"),
        label="design acceptance component sheet",
    )
    ws_component.append(["组件", "问题数量"])
    _apply_header_style(ws_component)

    # Placeholder: group by task_id as component proxy
    component_counts: dict[str, int] = {}
    for issue in issues:
        component = issue.get("task_id", "unknown")
        component_counts[component] = component_counts.get(component, 0) + 1

    for component, count in sorted(component_counts.items()):
        ws_component.append([component, count])

    _auto_adjust_column_width(ws_component)


def _build_competitor_template(wb: Workbook, issues: list[dict[str, Any]]) -> None:
    """Build the competitor template with 2 sheets."""
    # Sheet 1: 功能对比矩阵
    ws_matrix = _require_worksheet(wb.active, label="competitor matrix sheet")
    ws_matrix.title = "功能对比矩阵"

    headers = ["功能", "产品A", "产品B", "产品C"]
    ws_matrix.append(headers)
    _apply_header_style(ws_matrix)

    # Placeholder: use issue titles as features
    for issue in issues[:10]:  # Limit to 10 for demo
        row = [issue.get("title", ""), "✓", "✗", "✓"]
        ws_matrix.append(row)

    _auto_adjust_column_width(ws_matrix)

    # Sheet 2: 维度评分
    ws_score = _require_worksheet(wb.create_sheet("维度评分"), label="competitor score sheet")
    ws_score.append(["维度", "产品A", "产品B", "产品C"])
    _apply_header_style(ws_score)

    dimensions = ["易用性", "功能完整性", "性能", "视觉设计", "一致性"]
    for dimension in dimensions:
        ws_score.append([dimension, 8, 7, 9])

    _auto_adjust_column_width(ws_score)


def _render_html_report(
    issues: list[dict[str, Any]],
    *,
    target_path: Path,
    journey_map: Any,
    principles: Any,
) -> None:
    """Render a small standalone HTML report for Stage 7."""
    target_path.parent.mkdir(parents=True, exist_ok=True)
    issue_items = "\n".join(
        (
            "<li>"
            f"<strong>{html.escape(str(issue.get('id', '')))} {html.escape(str(issue.get('title', '')))}</strong>"
            f" <span>({html.escape(str(issue.get('severity', '')) or 'unknown')})</span><br/>"
            f"<span>{html.escape(str(issue.get('description', '')))}</span>"
            "</li>"
        )
        for issue in issues
    )
    payload = {
        "journey_map": journey_map,
        "principles": principles,
        "issues": issues,
    }
    document = (
        "<!DOCTYPE html>\n"
        "<html lang=\"zh-CN\">\n"
        "<head>\n"
        "<meta charset=\"utf-8\" />\n"
        "<title>UXEval HTML Report</title>\n"
        "<style>"
        "body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;margin:2rem;color:#1f2937;}"
        "h1,h2{margin:0 0 1rem;}"
        "section{margin:2rem 0;}"
        "ul{padding-left:1.25rem;}"
        "li{margin-bottom:0.75rem;}"
        "pre{background:#f3f4f6;padding:1rem;border-radius:8px;overflow:auto;}"
        "</style>\n"
        "</head>\n"
        "<body>\n"
        "<h1>UXEval Report</h1>\n"
        f"<p>Total issues: {len(issues)}</p>\n"
        "<section>\n"
        "<h2>Issue list</h2>\n"
        f"<ul>{issue_items}</ul>\n"
        "</section>\n"
        "<section>\n"
        "<h2>Structured payload</h2>\n"
        f"<pre>{html.escape(json.dumps(payload, ensure_ascii=False, indent=2, default=str))}</pre>\n"
        "</section>\n"
        "</body>\n"
        "</html>\n"
    )
    target_path.write_text(document, encoding="utf-8")


def _write_evidence_pack(
    evidence_dir: Path,
    *,
    issues: list[dict[str, Any]],
    journey_map: Any,
    principles: Any,
    excel_path: Path,
    html_path: Path,
    template: str,
) -> int:
    """Write a real evidence directory with a manifest and source payloads."""
    evidence_dir.mkdir(parents=True, exist_ok=True)
    manifest = {
        "generated_at": datetime.now(UTC).isoformat(),
        "template": template,
        "issue_count": len(issues),
        "issue_report_path": str(excel_path),
        "html_report_path": str(html_path),
    }
    files: dict[str, Any] = {
        "manifest.json": manifest,
        "issues.json": issues,
        "journey_map.json": journey_map,
        "principles.json": principles,
    }
    written = 0
    for filename, payload in files.items():
        (evidence_dir / filename).write_text(
            json.dumps(payload, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        written += 1
    return written


def _artifact_payload(
    *,
    artifact_id: str,
    output_type: str,
    path: Path,
    format: str,
    summary: str,
    **metadata: Any,
) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "id": artifact_id,
        "type": output_type,
        "path": str(path.resolve()),
        "format": format,
        "summary": summary,
    }
    payload.update(metadata)
    return payload


def _string_list(value: Any) -> list[str]:
    if isinstance(value, list):
        items: list[str] = []
        for item in value:
            cleaned = str(item).strip()
            if cleaned:
                items.append(cleaned)
        return items
    if isinstance(value, str):
        cleaned = value.strip()
        return [cleaned] if cleaned else []
    return []


def _unique_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    unique: list[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        unique.append(value)
    return unique


def _coerce_float(value: Any) -> float:
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return 0.0


def _coerce_int(value: Any) -> int:
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return 0


def _build_client_mode_benchmark_summary(
    *,
    output_dir: Path,
    bundle_dir: Path,
    issues: list[dict[str, Any]],
    demoted_main_issues: list[dict[str, Any]],
    key_unverified_issues: list[dict[str, Any]],
    evidence_assessment: dict[str, Any],
    audited_delivery_assessment: dict[str, Any],
) -> dict[str, Any]:
    existing_metrics = evidence_assessment.get("client_mode_metrics", {})
    if not isinstance(existing_metrics, dict):
        existing_metrics = {}
    existing_summary = evidence_assessment.get("benchmark_summary", {})
    if not isinstance(existing_summary, dict):
        existing_summary = {}
    coverage_summary = evidence_assessment.get("coverage_summary", {})
    if not isinstance(coverage_summary, dict):
        coverage_summary = {}
    audited_breakdown = audited_delivery_assessment.get("delivery_readiness_breakdown", {})
    if not isinstance(audited_breakdown, dict):
        audited_breakdown = {}

    existing_coverage = existing_metrics.get("coverage_metrics", {})
    if not isinstance(existing_coverage, dict):
        existing_coverage = {}
    existing_trust = existing_metrics.get("trust_metrics", {})
    if not isinstance(existing_trust, dict):
        existing_trust = {}
    existing_success = existing_metrics.get("success_metrics", {})
    if not isinstance(existing_success, dict):
        existing_success = {}
    existing_human = existing_metrics.get("human_burden_metrics", {})
    if not isinstance(existing_human, dict):
        existing_human = {}

    delivery_status = str(audited_delivery_assessment.get("delivery_status", "supplement_required"))
    planned_page_count = max(1, _coerce_int(coverage_summary.get("planned_page_count", 0)))
    trusted_mapping_count = _coerce_int(coverage_summary.get("final_delivery_trusted_mapping_count", 0))
    provisional_mapping_count = _coerce_int(coverage_summary.get("fusion_provisional_mapping_count", 0))
    conflict_count = _coerce_int(coverage_summary.get("fusion_conflict_count", 0))
    clarification_count = _coerce_int(coverage_summary.get("clarification_needed_count", 0))
    unverified_leakage_rate = round(
        (len(demoted_main_issues) + len(key_unverified_issues)) / max(1, len(issues)),
        3,
    )

    coverage_metrics = {
        "critical_path_page_hit_rate": _coerce_float(
            existing_coverage.get(
                "critical_path_page_hit_rate",
                coverage_summary.get("final_delivery_page_coverage_ratio", 0.0),
            )
        ),
        "critical_path_state_hit_rate": _coerce_float(
            existing_coverage.get(
                "critical_path_state_hit_rate",
                coverage_summary.get("final_delivery_state_coverage_ratio", 0.0),
            )
        ),
        "p0_page_coverage": _coerce_float(existing_coverage.get("p0_page_coverage", 0.0)),
        "p0_state_coverage": _coerce_float(existing_coverage.get("p0_state_coverage", 0.0)),
        "p1_page_coverage": _coerce_float(existing_coverage.get("p1_page_coverage", 0.0)),
        "p1_state_coverage": _coerce_float(existing_coverage.get("p1_state_coverage", 0.0)),
    }
    trust_metrics = {
        "trusted_mapping_rate": _coerce_float(
            existing_trust.get(
                "trusted_mapping_rate",
                round(trusted_mapping_count / planned_page_count, 3),
            )
        ),
        "provisional_mapping_rate": _coerce_float(
            existing_trust.get(
                "provisional_mapping_rate",
                round(min(1.0, provisional_mapping_count / planned_page_count), 3),
            )
        ),
        "conflicting_mapping_rate": _coerce_float(
            existing_trust.get(
                "conflicting_mapping_rate",
                round(min(1.0, conflict_count / planned_page_count), 3),
            )
        ),
        "unverified_leakage_rate": unverified_leakage_rate,
    }
    success_metrics = {
        "final_delivery_ready_rate": 1.0 if delivery_status == "final_delivery_ready" else 0.0,
        "fallback_safe_rate": 1.0 if delivery_status == "fallback_safe" else 0.0,
        "supplement_required_rate": 1.0 if delivery_status == "supplement_required" else 0.0,
        "blocked_rate": 1.0 if delivery_status == "blocked" else 0.0,
        "first_pass_final_rate": _coerce_float(existing_success.get("first_pass_final_rate", 0.0)),
        "auto_remediation_lift": _coerce_float(existing_success.get("auto_remediation_lift", 0.0)),
        "salvageable_input_rate": _coerce_float(existing_success.get("salvageable_input_rate", 0.0)),
    }
    human_burden_metrics = {
        "clarification_item_count": _coerce_int(
            existing_human.get("clarification_item_count", clarification_count)
        ),
        "supplement_request_precision": _coerce_float(
            existing_human.get("supplement_request_precision", 0.0)
        ),
        "low_value_work_return_rate": _coerce_float(
            existing_human.get("low_value_work_return_rate", 0.0)
        ),
    }

    failing_final_gates = _string_list(audited_breakdown.get("failing_final_gates"))
    met_metrics = _unique_preserve_order(
        [
            *(
                ["critical_path_page_hit_rate>=0.90"]
                if coverage_metrics["critical_path_page_hit_rate"] >= 0.9
                else []
            ),
            *(
                ["critical_path_state_hit_rate>=0.90"]
                if coverage_metrics["critical_path_state_hit_rate"] >= 0.9
                else []
            ),
            *(
                ["trusted_mapping_rate>=0.90"]
                if trust_metrics["trusted_mapping_rate"] >= 0.9
                else []
            ),
            *(
                ["unverified_leakage_rate==0.00"]
                if trust_metrics["unverified_leakage_rate"] == 0.0
                else []
            ),
            *(
                ["clarification_item_count==0"]
                if human_burden_metrics["clarification_item_count"] == 0
                else []
            ),
        ]
    )
    unmet_metrics = _unique_preserve_order(
        [
            *(
                ["critical_path_page_hit_rate<0.90"]
                if coverage_metrics["critical_path_page_hit_rate"] < 0.9
                else []
            ),
            *(
                ["critical_path_state_hit_rate<0.90"]
                if coverage_metrics["critical_path_state_hit_rate"] < 0.9
                else []
            ),
            *(
                ["trusted_mapping_rate<0.90"]
                if trust_metrics["trusted_mapping_rate"] < 0.9
                else []
            ),
            *(
                ["unverified_leakage_rate>0.00"]
                if trust_metrics["unverified_leakage_rate"] > 0.0
                else []
            ),
            *(
                ["clarification_item_count>0"]
                if human_burden_metrics["clarification_item_count"] > 0
                else []
            ),
            *[f"gate:{gate}" for gate in failing_final_gates],
        ]
    )

    distance_to_90_plus = _unique_preserve_order(
        [
            *(
                [
                    "critical_path_page_hit_rate still below 0.90"
                ]
                if coverage_metrics["critical_path_page_hit_rate"] < 0.9
                else []
            ),
            *(
                [
                    "critical_path_state_hit_rate still below 0.90"
                ]
                if coverage_metrics["critical_path_state_hit_rate"] < 0.9
                else []
            ),
            *(
                [
                    "trusted_mapping_rate still below 0.90"
                ]
                if trust_metrics["trusted_mapping_rate"] < 0.9
                else []
            ),
            *(
                [
                    f"unverified issue leakage still at {trust_metrics['unverified_leakage_rate']:.3f}"
                ]
                if trust_metrics["unverified_leakage_rate"] > 0.0
                else []
            ),
            *[f"final gate still fails at {gate}" for gate in failing_final_gates],
        ]
    )

    artifact_paths = {
        "benchmark_dir": str((output_dir / _BENCHMARK_DIRNAME).resolve()),
        "benchmark_json_path": str((output_dir / _BENCHMARK_DIRNAME / _BENCHMARK_JSON_FILENAME).resolve()),
        "benchmark_markdown_path": str((output_dir / _BENCHMARK_DIRNAME / _BENCHMARK_MD_FILENAME).resolve()),
        "delivery_audit_bundle_path": str(bundle_dir.resolve()),
        "bounded_issue_pass_path": str((bundle_dir / "bounded_issue_pass.md").resolve()),
        "supplement_request_path": str((bundle_dir / "supplement_request.md").resolve()),
        "audited_delivery_assessment_path": str((bundle_dir / "audited_delivery_assessment.json").resolve()),
    }
    root_cause = str(existing_summary.get("root_cause", "unknown")).strip() or "unknown"
    if root_cause == "unknown":
        if delivery_status in {"supplement_required", "blocked"}:
            root_cause = "input_truly_insufficient"
        elif delivery_status == "fallback_safe" and trust_metrics["unverified_leakage_rate"] > 0.0:
            root_cause = "system_ingestion_gap"
        else:
            root_cause = "mixed_gap_profile"

    summary_headline = {
        "final_delivery_ready": "当前 run 已通过 final delivery audit，可进入最终交付。",
        "fallback_safe": "当前 run 只达到 bounded fallback；audit 已阻止它冒充 final。",
        "supplement_required": "当前 run 仍需补证据后再尝试升级为 final。",
        "blocked": "当前 run 被阻断，audit 认定现有证据不适合继续交付。",
    }.get(delivery_status, "当前 run 仍需进一步诊断。")

    return {
        "contract_version": _BENCHMARK_CONTRACT_VERSION,
        "run_mode": "client",
        "input_quality_class": str(
            existing_summary.get("input_quality_class")
            or existing_metrics.get("input_quality_class")
            or "unknown"
        ),
        "delivery_status": delivery_status,
        "metrics": {
            "contract_version": _BENCHMARK_CONTRACT_VERSION,
            "run_mode": "client",
            "input_quality_class": str(
                existing_summary.get("input_quality_class")
                or existing_metrics.get("input_quality_class")
                or "unknown"
            ),
            "delivery_status": delivery_status,
            "coverage_metrics": coverage_metrics,
            "trust_metrics": trust_metrics,
            "success_metrics": success_metrics,
            "human_burden_metrics": human_burden_metrics,
            "metric_notes": {
                "unverified_leakage_rate": "overwritten by delivery audit using the actual qualified vs unverified issue split",
            },
        },
        "met_metrics": met_metrics,
        "unmet_metrics": unmet_metrics,
        "distance_to_90_plus": distance_to_90_plus,
        "root_cause": root_cause,
        "next_best_action": audited_breakdown.get("next_best_action", ""),
        "summary_headline": summary_headline,
        "artifact_paths": artifact_paths,
    }


def _render_client_mode_benchmark_summary(summary: dict[str, Any]) -> str:
    metrics = summary.get("metrics", {})
    coverage = metrics.get("coverage_metrics", {}) if isinstance(metrics, dict) else {}
    trust = metrics.get("trust_metrics", {}) if isinstance(metrics, dict) else {}
    success = metrics.get("success_metrics", {}) if isinstance(metrics, dict) else {}
    burden = metrics.get("human_burden_metrics", {}) if isinstance(metrics, dict) else {}
    lines = [
        "# Client Mode Benchmark Summary",
        "",
        f"- contract_version: {summary.get('contract_version', '')}",
        f"- run_mode: {summary.get('run_mode', '')}",
        f"- input_quality_class: {summary.get('input_quality_class', '')}",
        f"- delivery_status: {summary.get('delivery_status', '')}",
        f"- summary_headline: {summary.get('summary_headline', '')}",
        f"- root_cause: {summary.get('root_cause', '')}",
        f"- next_best_action: {summary.get('next_best_action', '')}",
        "",
        "## Coverage metrics",
        f"- critical_path_page_hit_rate: {_coerce_float(coverage.get('critical_path_page_hit_rate', 0.0)):.3f}",
        f"- critical_path_state_hit_rate: {_coerce_float(coverage.get('critical_path_state_hit_rate', 0.0)):.3f}",
        f"- p0_page_coverage: {_coerce_float(coverage.get('p0_page_coverage', 0.0)):.3f}",
        f"- p0_state_coverage: {_coerce_float(coverage.get('p0_state_coverage', 0.0)):.3f}",
        f"- p1_page_coverage: {_coerce_float(coverage.get('p1_page_coverage', 0.0)):.3f}",
        f"- p1_state_coverage: {_coerce_float(coverage.get('p1_state_coverage', 0.0)):.3f}",
        "",
        "## Trust metrics",
        f"- trusted_mapping_rate: {_coerce_float(trust.get('trusted_mapping_rate', 0.0)):.3f}",
        f"- provisional_mapping_rate: {_coerce_float(trust.get('provisional_mapping_rate', 0.0)):.3f}",
        f"- conflicting_mapping_rate: {_coerce_float(trust.get('conflicting_mapping_rate', 0.0)):.3f}",
        f"- unverified_leakage_rate: {_coerce_float(trust.get('unverified_leakage_rate', 0.0)):.3f}",
        "",
        "## Success / throughput metrics",
        f"- final_delivery_ready_rate: {_coerce_float(success.get('final_delivery_ready_rate', 0.0)):.3f}",
        f"- fallback_safe_rate: {_coerce_float(success.get('fallback_safe_rate', 0.0)):.3f}",
        f"- supplement_required_rate: {_coerce_float(success.get('supplement_required_rate', 0.0)):.3f}",
        f"- blocked_rate: {_coerce_float(success.get('blocked_rate', 0.0)):.3f}",
        f"- first_pass_final_rate: {_coerce_float(success.get('first_pass_final_rate', 0.0)):.3f}",
        f"- auto_remediation_lift: {_coerce_float(success.get('auto_remediation_lift', 0.0)):.3f}",
        f"- salvageable_input_rate: {_coerce_float(success.get('salvageable_input_rate', 0.0)):.3f}",
        "",
        "## Human burden metrics",
        f"- clarification_item_count: {_coerce_int(burden.get('clarification_item_count', 0))}",
        f"- supplement_request_precision: {_coerce_float(burden.get('supplement_request_precision', 0.0)):.3f}",
        f"- low_value_work_return_rate: {_coerce_float(burden.get('low_value_work_return_rate', 0.0)):.3f}",
        "",
        "## Met metrics",
    ]
    met_metrics = _string_list(summary.get("met_metrics"))
    lines.extend([f"- {item}" for item in met_metrics] or ["- 无"])
    lines.extend(["", "## Unmet metrics"])
    unmet_metrics = _string_list(summary.get("unmet_metrics"))
    lines.extend([f"- {item}" for item in unmet_metrics] or ["- 无"])
    lines.extend(["", "## Distance to 90%+"])
    distance = _string_list(summary.get("distance_to_90_plus"))
    lines.extend([f"- {item}" for item in distance] or ["- 当前已达到或超过关键 90%+ 目标。"])
    return "\n".join(lines).strip() + "\n"


def _write_client_mode_benchmark_summary(
    *,
    output_dir: Path,
    summary: dict[str, Any],
) -> None:
    benchmark_dir = output_dir / _BENCHMARK_DIRNAME
    benchmark_dir.mkdir(parents=True, exist_ok=True)
    (benchmark_dir / _BENCHMARK_JSON_FILENAME).write_text(
        json.dumps(summary, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    (benchmark_dir / _BENCHMARK_MD_FILENAME).write_text(
        _render_client_mode_benchmark_summary(summary),
        encoding="utf-8",
    )


def _issue_audit_failures(issue: dict[str, Any]) -> list[str]:
    failures: list[str] = []
    if str(issue.get("verification_status", "")).strip() != "verified":
        failures.append("verification_status != verified")
    if not _string_list(issue.get("evidence_refs")):
        failures.append("missing evidence_refs")
    if not _string_list(issue.get("evidence_basis")):
        failures.append("missing evidence_basis")
    return failures


def _is_key_unverified_issue(issue: dict[str, Any]) -> bool:
    severity = str(issue.get("severity", "")).strip().lower()
    if severity in {"critical", "major"}:
        return True
    if issue.get("blocks_final_delivery") is True:
        return True
    return severity == ""


def _gate_metric_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return []


def _render_bounded_issue_pass(
    *,
    safe_main_issues: list[dict[str, Any]],
    demoted_main_issues: list[dict[str, Any]],
    audited_delivery_assessment: dict[str, Any],
) -> str:
    lines = [
        "# Bounded Issue Pass",
        "",
        f"- delivery_status: {audited_delivery_assessment['delivery_status']}",
        f"- qualified_issue_count: {len(safe_main_issues)}",
        f"- demoted_main_issue_count: {len(demoted_main_issues)}",
        "",
        "## Capture Mission pass lines",
    ]
    for line in _string_list(audited_delivery_assessment.get("fallback_pass_line")):
        lines.append(f"- fallback: {line}")
    for line in _string_list(audited_delivery_assessment.get("final_delivery_pass_line")):
        lines.append(f"- final: {line}")
    breakdown = audited_delivery_assessment.get("delivery_readiness_breakdown", {})
    if isinstance(breakdown, dict):
        lines.extend(
            [
                "",
                "## Delivery readiness breakdown",
                f"- final_gate_pass: {breakdown.get('final_gate_pass', False)}",
                f"- fallback_gate_pass: {breakdown.get('fallback_gate_pass', False)}",
            ]
        )
        failing_final = _string_list(breakdown.get("failing_final_gates"))
        failing_fallback = _string_list(breakdown.get("failing_fallback_gates"))
        if failing_final:
            lines.append(f"- failing_final_gates: {' / '.join(failing_final)}")
        if failing_fallback:
            lines.append(f"- failing_fallback_gates: {' / '.join(failing_fallback)}")
        next_best_action = str(breakdown.get("next_best_action", "")).strip()
        if next_best_action:
            lines.append(f"- next_best_action: {next_best_action}")
    lines.extend(
        [
            "",
        "## 当前可直接使用的可信主结论",
        ]
    )
    if safe_main_issues:
        for issue in safe_main_issues:
            evidence_refs = ", ".join(_string_list(issue.get("evidence_refs")))
            evidence_basis = "；".join(_string_list(issue.get("evidence_basis"))[:3])
            lines.extend(
                [
                    f"### {issue.get('id', '')} {issue.get('title', '')}",
                    f"- severity: {issue.get('severity', '')}",
                    f"- user_impact: {issue.get('user_impact', '')}",
                    f"- evidence_refs: {evidence_refs or '(missing)'}",
                    f"- evidence_basis: {evidence_basis or '(missing)'}",
                    f"- suggestion: {issue.get('suggestion', '')}",
                    "",
                ]
            )
    else:
        lines.extend(
            [
                "- 当前没有足够稳定的主问题结论可直接分享。",
                "",
            ]
        )

    lines.append("## 被 runtime audit 从主清单剔除的问题")
    if demoted_main_issues:
        for issue in demoted_main_issues:
            lines.extend(
                [
                    f"- {issue.get('id', '')} {issue.get('title', '')}: "
                    + "；".join(_string_list(issue.get("audit_failures"))),
                ]
            )
    else:
        lines.append("- 无")

    return "\n".join(lines).strip() + "\n"


def _render_supplement_request(
    *,
    audited_delivery_assessment: dict[str, Any],
    evidence_assessment: dict[str, Any],
    key_unverified_issues: list[dict[str, Any]],
    demoted_main_issues: list[dict[str, Any]],
) -> str:
    coverage_summary = evidence_assessment.get("coverage_summary", {})
    if not isinstance(coverage_summary, dict):
        coverage_summary = {}

    lines = [
        "# Supplement Request",
        "",
        f"- audited_delivery_status: {audited_delivery_assessment['delivery_status']}",
        f"- status_reason: {audited_delivery_assessment.get('status_reason', '')}",
        "",
        "## Required actions",
    ]
    actions = _string_list(audited_delivery_assessment.get("required_actions"))
    if actions:
        lines.extend(f"- {action}" for action in actions)
    else:
        lines.append("- 无额外补料要求。")

    missing_tasks = _string_list(coverage_summary.get("missing_tasks"))
    missing_states = _string_list(coverage_summary.get("missing_state_categories"))
    low_readability = _string_list(coverage_summary.get("low_readability_paths"))
    missing_desc = _string_list(coverage_summary.get("missing_description_paths"))
    missing_ocr = _string_list(coverage_summary.get("missing_ocr_paths"))
    missing_planned_pages = _string_list(coverage_summary.get("missing_critical_pages"))
    missing_planned_states = _string_list(coverage_summary.get("missing_planned_states"))
    missing_required_descriptions = _string_list(coverage_summary.get("missing_required_description_pages"))
    naming_issues = _string_list(coverage_summary.get("naming_issues"))
    fusion_conflicting_paths = _string_list(coverage_summary.get("fusion_conflicting_paths"))
    fusion_unresolved_paths = _string_list(coverage_summary.get("fusion_unresolved_paths"))
    critical_path_final_blockers = _string_list(coverage_summary.get("final_delivery_missing_critical_paths"))
    critical_path_fallback_blockers = _string_list(coverage_summary.get("fallback_missing_critical_paths"))

    lines.extend(
        [
            "",
            "## Missing evidence coverage",
        ]
    )
    coverage_items = [
        ("missing_tasks", missing_tasks),
        ("missing_state_categories", missing_states),
        ("missing_critical_pages", missing_planned_pages),
        ("missing_planned_states", missing_planned_states),
        ("missing_required_description_pages", missing_required_descriptions),
        ("low_readability_paths", low_readability),
        ("missing_description_paths", missing_desc),
        ("missing_ocr_paths", missing_ocr),
        ("naming_issues", naming_issues),
        ("fusion_conflicting_paths", fusion_conflicting_paths),
        ("fusion_unresolved_paths", fusion_unresolved_paths),
        ("final_delivery_missing_critical_paths", critical_path_final_blockers),
        ("fallback_missing_critical_paths", critical_path_fallback_blockers),
    ]
    any_coverage = False
    for label, values in coverage_items:
        if not values:
            continue
        any_coverage = True
        lines.append(f"- {label}: {'；'.join(values[:6])}")
    if not any_coverage:
        lines.append("- 当前没有额外的结构化 coverage gap。")

    lines.extend(["", "## Capture Mission pass lines"])
    final_pass_line = _string_list(audited_delivery_assessment.get("final_delivery_pass_line"))
    fallback_pass_line = _string_list(audited_delivery_assessment.get("fallback_pass_line"))
    if final_pass_line:
        lines.extend(f"- final: {line}" for line in final_pass_line)
    if fallback_pass_line:
        lines.extend(f"- fallback: {line}" for line in fallback_pass_line)
    if not final_pass_line and not fallback_pass_line:
        lines.append("- 当前未附带 Capture Mission pass line。")

    breakdown = audited_delivery_assessment.get("delivery_readiness_breakdown", {})
    if isinstance(breakdown, dict):
        lines.extend(["", "## Delivery readiness breakdown"])
        lines.append(f"- final_gate_pass: {breakdown.get('final_gate_pass', False)}")
        lines.append(f"- fallback_gate_pass: {breakdown.get('fallback_gate_pass', False)}")
        for gate in breakdown.get("gates", []):
            if not isinstance(gate, dict):
                continue
            gate_name = str(gate.get("gate", "")).strip()
            final_status = str(gate.get("final_status", "")).strip()
            fallback_status = str(gate.get("fallback_status", "")).strip()
            failure_reasons = _string_list(gate.get("failure_reasons"))
            if not gate_name:
                continue
            lines.append(f"- {gate_name}: final={final_status} fallback={fallback_status}")
            for reason in failure_reasons[:3]:
                lines.append(f"  - {reason}")
        next_best_action = str(breakdown.get("next_best_action", "")).strip()
        if next_best_action:
            lines.append(f"- next_best_action: {next_best_action}")

    lines.extend(["", "## Still-unverified issues"])
    if key_unverified_issues or demoted_main_issues:
        for issue in [*key_unverified_issues, *demoted_main_issues]:
            issue_id = issue.get("id", "")
            title = issue.get("title", "")
            blockers = _string_list(issue.get("blocked_by")) + _string_list(issue.get("audit_failures"))
            lines.append(f"- {issue_id} {title}: {'；'.join(blockers) or '待补充证据'}")
    else:
        lines.append("- 无")

    return "\n".join(lines).strip() + "\n"


def audit_delivery_readiness(
    issues: list[dict[str, Any]],
    unverified_issues: list[dict[str, Any]] | None = None,
    evidence_assessment: dict[str, Any] | None = None,
    delivery_assessment: dict[str, Any] | None = None,
    capture_mission: dict[str, Any] | None = None,
    output_dir: str | None = None,
    run_id: str | None = None,
    skill_name: str | None = None,
    skill_version: str | None = None,
    stage_id: str | None = None,
) -> dict[str, Any]:
    """Deterministically audit final-delivery readiness and write a bounded package."""
    unverified_source = list(unverified_issues or [])
    evidence = dict(evidence_assessment or {})
    llm_delivery = dict(delivery_assessment or {})

    if output_dir is None:
        raise ExcelBuilderError("output_dir is required for delivery audit packaging")
    bundle_dir = Path(output_dir) / "delivery_audit_bundle"
    bundle_dir.mkdir(parents=True, exist_ok=True)

    safe_main_issues: list[dict[str, Any]] = []
    demoted_main_issues: list[dict[str, Any]] = []
    demoted_actions: list[str] = []
    for issue in issues:
        failures = _issue_audit_failures(issue)
        if failures:
            demoted = dict(issue)
            demoted["audit_failures"] = failures
            demoted["demoted_from_main_list"] = True
            demoted_main_issues.append(demoted)
            demoted_actions.append(
                f"修正主清单问题 {issue.get('id', '')} 的审计缺口：{'；'.join(failures)}"
            )
        else:
            safe_main_issues.append(issue)

    merged_unverified = [*unverified_source, *demoted_main_issues]
    key_unverified = [issue for issue in merged_unverified if _is_key_unverified_issue(issue)]

    evidence_delivery_status = str(evidence.get("delivery_status", "blocked")).strip() or "blocked"
    missing_coverage = _string_list(evidence.get("missing_coverage"))
    evidence_required_actions = _string_list(evidence.get("required_actions"))
    evidence_verification_gaps = _string_list(evidence.get("verification_gaps"))
    coverage_summary = evidence.get("coverage_summary", {})
    if not isinstance(coverage_summary, dict):
        coverage_summary = {}
    mission = capture_mission if isinstance(capture_mission, dict) else {}
    mission_version = str(
        mission.get("mission_version")
        or coverage_summary.get("capture_mission_version")
        or ""
    ).strip()
    mission_path = str(
        mission.get("mission_doc_path")
        or coverage_summary.get("capture_mission_path")
        or ""
    ).strip()
    critical_flows = _string_list(
        mission.get("critical_flows") or coverage_summary.get("critical_flows")
    )
    critical_path_coverage = evidence.get("critical_path_coverage_summary", {})
    if not isinstance(critical_path_coverage, dict):
        critical_path_coverage = {}
    capture_order = _string_list(
        mission.get("capture_order") or coverage_summary.get("capture_order")
    )
    final_delivery_pass_line = _string_list(
        mission.get("final_delivery_pass_line") or coverage_summary.get("final_delivery_pass_line")
    )
    fallback_pass_line = _string_list(
        mission.get("fallback_pass_line") or coverage_summary.get("fallback_pass_line")
    )
    fusion_summary = evidence.get("fusion_summary", {})
    if not isinstance(fusion_summary, dict):
        fusion_summary = {}
    trusted_page_mappings = fusion_summary.get("trusted_page_mappings", [])
    provisional_mappings = fusion_summary.get("provisional_mappings", [])
    conflicting_groups = fusion_summary.get("conflicting_evidence_groups", [])
    unresolved_ambiguities = fusion_summary.get("unresolved_ambiguities", [])
    if not isinstance(trusted_page_mappings, list):
        trusted_page_mappings = []
    if not isinstance(provisional_mappings, list):
        provisional_mappings = []
    if not isinstance(conflicting_groups, list):
        conflicting_groups = []
    if not isinstance(unresolved_ambiguities, list):
        unresolved_ambiguities = []
    critical_path_records = critical_path_coverage.get("critical_paths", [])
    if not isinstance(critical_path_records, list):
        critical_path_records = []
    evidence_breakdown = evidence.get("delivery_readiness_breakdown")
    if not isinstance(evidence_breakdown, dict):
        evidence_breakdown = coverage_summary.get("delivery_readiness_breakdown", {})
    if not isinstance(evidence_breakdown, dict):
        evidence_breakdown = {}

    audit_failures: list[str] = []
    final_critical_path_failures = _string_list(coverage_summary.get("final_delivery_missing_critical_paths"))
    fallback_critical_path_failures = _string_list(coverage_summary.get("fallback_missing_critical_paths"))
    clarification_needed_count = int(coverage_summary.get("clarification_needed_count", 0) or 0)
    clarification_unlocks_final_count = int(
        coverage_summary.get("clarification_unlocks_final_count", 0) or 0
    )
    unresolved_ambiguity_count = int(
        coverage_summary.get("fusion_unresolved_ambiguity_count", 0) or 0
    )

    critical_path_gate = {
        "gate": "critical_path_coverage",
        "label": "P0/P1 critical path coverage",
        "required_for_final": True,
        "required_for_fallback": True,
        "final_status": "fail" if final_critical_path_failures else "pass",
        "fallback_status": "fail" if fallback_critical_path_failures else "pass",
        "failure_reasons": _unique_preserve_order(
            [
                *(
                    [
                        "critical business path final gate not met: "
                        + "；".join(final_critical_path_failures[:4])
                    ]
                    if final_critical_path_failures
                    else []
                ),
                *(
                    [
                        "P0 critical path still blocks bounded fallback: "
                        + "；".join(fallback_critical_path_failures[:4])
                    ]
                    if fallback_critical_path_failures
                    else []
                ),
            ]
        ),
        "next_actions": [
            action
            for action in evidence_required_actions
            if "critical path" in action or "P0" in action
        ][:3],
        "metrics": {
            "failing_final_paths": final_critical_path_failures,
            "failing_fallback_paths": fallback_critical_path_failures,
            "critical_path_count": len(critical_path_records),
        },
    }

    trusted_evidence_final_failures = _string_list(
        evidence_breakdown.get("failing_final_gates")
    )
    if "trusted_evidence_sufficiency" in trusted_evidence_final_failures:
        trusted_evidence_final_failures = []
    if evidence_delivery_status != "final_delivery_ready":
        trusted_evidence_final_failures.append(
            f"evidence_assessment.delivery_status={evidence_delivery_status}, not final_delivery_ready"
        )
    if missing_coverage:
        trusted_evidence_final_failures.append("critical evidence coverage gaps remain unresolved")
    if _string_list(coverage_summary.get("final_delivery_missing_critical_pages")):
        trusted_evidence_final_failures.append("Capture Mission final pass line not met: missing critical pages")
    if _string_list(coverage_summary.get("final_delivery_missing_planned_states")):
        trusted_evidence_final_failures.append("Capture Mission final pass line not met: missing critical states")
    if _string_list(coverage_summary.get("final_delivery_missing_required_description_pages")):
        trusted_evidence_final_failures.append("Capture Mission final pass line not met: missing required descriptions")

    trusted_evidence_fallback_failures: list[str] = []
    if evidence_delivery_status not in {"final_delivery_ready", "fallback_safe"}:
        trusted_evidence_fallback_failures.append(
            f"evidence_assessment.delivery_status={evidence_delivery_status}, below fallback_safe"
        )
    if fallback_critical_path_failures:
        trusted_evidence_fallback_failures.append("fallback critical path pass line not met")

    trusted_evidence_gate = {
        "gate": "trusted_evidence_sufficiency",
        "label": "Trusted evidence sufficiency",
        "required_for_final": True,
        "required_for_fallback": True,
        "final_status": "fail" if trusted_evidence_final_failures else "pass",
        "fallback_status": "fail" if trusted_evidence_fallback_failures else "pass",
        "failure_reasons": _unique_preserve_order(
            trusted_evidence_final_failures + trusted_evidence_fallback_failures
        ),
        "next_actions": [
            action
            for action in evidence_required_actions
            if "截图" in action or "说明" in action or "状态" in action or "页面" in action
        ][:4],
        "metrics": {
            "final_delivery_page_coverage_ratio": coverage_summary.get("final_delivery_page_coverage_ratio"),
            "final_delivery_state_coverage_ratio": coverage_summary.get("final_delivery_state_coverage_ratio"),
            "final_delivery_trusted_mapping_count": coverage_summary.get("final_delivery_trusted_mapping_count"),
            "readable_ratio": coverage_summary.get("readable_ratio"),
            "text_rich_ratio": coverage_summary.get("text_rich_ratio"),
        },
    }

    clarification_failures: list[str] = []
    if clarification_needed_count > 0:
        clarification_failures.append(
            f"{clarification_needed_count} clarification item(s) still unresolved"
        )
    if clarification_unlocks_final_count > 0:
        clarification_failures.append(
            f"{clarification_unlocks_final_count} clarification item(s) still block final delivery"
        )
    if unresolved_ambiguity_count > 0:
        clarification_failures.append(
            f"{unresolved_ambiguity_count} unresolved ambiguity group(s) still remain"
        )
    clarification_gate = {
        "gate": "clarification_residue",
        "label": "Clarification residue",
        "required_for_final": True,
        "required_for_fallback": False,
        "final_status": "fail" if clarification_failures else "pass",
        "fallback_status": "pass",
        "failure_reasons": clarification_failures,
        "next_actions": [
            action
            for action in evidence_required_actions
            if "歧义截图" in action or "确认" in action
        ][:2],
        "metrics": {
            "clarification_needed_count": clarification_needed_count,
            "clarification_unlocks_final_count": clarification_unlocks_final_count,
            "unresolved_ambiguity_count": unresolved_ambiguity_count,
            "clarification_paths": _string_list(coverage_summary.get("clarification_relative_paths")),
        },
    }

    issue_qualification_failures: list[str] = []
    if not safe_main_issues:
        issue_qualification_failures.append("no qualified issue left for the final main list")
    if demoted_main_issues:
        issue_qualification_failures.append(
            f"{len(demoted_main_issues)} main-list issue(s) failed deterministic verification"
        )
    if key_unverified:
        issue_qualification_failures.append(
            f"{len(key_unverified)} key unverified issue(s) still affect the main conclusion"
        )
    issue_qualification_gate = {
        "gate": "issue_qualification",
        "label": "Issue-level qualification",
        "required_for_final": True,
        "required_for_fallback": True,
        "final_status": "fail" if issue_qualification_failures else "pass",
        "fallback_status": "pass" if safe_main_issues else "fail",
        "failure_reasons": issue_qualification_failures,
        "next_actions": demoted_actions[:4],
        "metrics": {
            "main_issue_count": len(issues),
            "qualified_issue_count": len(safe_main_issues),
            "demoted_main_issue_count": len(demoted_main_issues),
            "key_unverified_issue_count": len(key_unverified),
        },
    }

    gates = [
        critical_path_gate,
        trusted_evidence_gate,
        clarification_gate,
        issue_qualification_gate,
    ]
    failing_final_gates = [
        str(gate["gate"])
        for gate in gates
        if gate["required_for_final"] and gate["final_status"] != "pass"
    ]
    failing_fallback_gates = [
        str(gate["gate"])
        for gate in gates
        if gate["required_for_fallback"] and gate["fallback_status"] != "pass"
    ]

    final_delivery_ready = not failing_final_gates
    fallback_gate_pass = not failing_fallback_gates and bool(safe_main_issues)
    if final_delivery_ready:
        audited_status = "final_delivery_ready"
    elif fallback_gate_pass:
        audited_status = "fallback_safe"
    elif evidence_delivery_status == "blocked":
        audited_status = "blocked"
    else:
        audited_status = "supplement_required"
    fallback_safe = audited_status == "fallback_safe"

    gate_failure_reasons = _unique_preserve_order(
        [
            reason
            for gate in gates
            for reason in _string_list(gate.get("failure_reasons"))
        ]
    )
    audit_failures.extend(gate_failure_reasons)

    required_actions = _unique_preserve_order(
        evidence_required_actions + demoted_actions
    )
    if audited_status == "fallback_safe" and not required_actions:
        required_actions.append("补关键页面或关键状态证据后再升级为最终报告")

    verification_gaps = _unique_preserve_order(
        evidence_verification_gaps
        + _string_list(evidence.get("missing_coverage"))
        + [
            f"{issue.get('id', '')} requires verification before entering the final main issue list"
            for issue in demoted_main_issues
        ]
    )

    if final_delivery_ready:
        status_reason = (
            "Runtime delivery audit passed: all main-list issues are verified, evidence-complete, "
            "and no blocking coverage gap remains."
        )
    elif audited_status == "fallback_safe":
        status_reason = (
            "Runtime delivery audit downgraded this run to bounded fallback sharing. "
            "Some verified conclusions are usable, but final report release is blocked."
        )
    elif audited_status == "blocked":
        status_reason = (
            "Runtime delivery audit blocked delivery because evidence is not trustworthy enough "
            "for either final report or bounded sharing."
        )
    else:
        status_reason = (
            "Runtime delivery audit requires supplementary evidence before this run can become "
            "a reliable final delivery."
        )

    audited_delivery_assessment: dict[str, Any] = {
        "delivery_status": audited_status,
        "final_delivery_ready": final_delivery_ready,
        "fallback_safe": fallback_safe,
        "status_reason": status_reason,
        "confidence": "high" if final_delivery_ready else "medium" if fallback_safe else "low",
        "required_actions": required_actions,
        "verification_gaps": verification_gaps,
        "evidence_basis": [
            f"main_issue_count={len(issues)}",
            f"qualified_issue_count={len(safe_main_issues)}",
            f"demoted_main_issue_count={len(demoted_main_issues)}",
            f"unverified_issue_count={len(merged_unverified)}",
            f"evidence_delivery_status={evidence_delivery_status}",
            f"llm_delivery_status={str(llm_delivery.get('delivery_status', 'unknown'))}",
        ],
        "audit_failures": audit_failures,
        "coverage_summary": {
            "missing_tasks": _string_list(coverage_summary.get("missing_tasks")),
            "missing_state_categories": _string_list(coverage_summary.get("missing_state_categories")),
            "missing_critical_pages": _string_list(coverage_summary.get("missing_critical_pages")),
            "missing_planned_states": _string_list(coverage_summary.get("missing_planned_states")),
            "missing_required_description_pages": _string_list(coverage_summary.get("missing_required_description_pages")),
            "final_delivery_missing_critical_paths": final_critical_path_failures,
            "fallback_missing_critical_paths": fallback_critical_path_failures,
            "naming_issues": _string_list(coverage_summary.get("naming_issues")),
            "low_readability_paths": _string_list(coverage_summary.get("low_readability_paths")),
            "missing_description_paths": _string_list(coverage_summary.get("missing_description_paths")),
            "missing_ocr_paths": _string_list(coverage_summary.get("missing_ocr_paths")),
            "fusion_conflicting_paths": [
                str(item.get("relative_path", "")).strip()
                for item in conflicting_groups
                if isinstance(item, dict) and str(item.get("relative_path", "")).strip()
            ],
            "fusion_unresolved_paths": [
                str(item.get("relative_path", "")).strip()
                for item in unresolved_ambiguities
                if isinstance(item, dict) and str(item.get("relative_path", "")).strip()
            ],
        },
        "capture_mission_version": mission_version or None,
        "capture_mission_path": mission_path or None,
        "critical_flows": critical_flows,
        "capture_order": capture_order,
        "final_delivery_pass_line": final_delivery_pass_line,
        "fallback_pass_line": fallback_pass_line,
        "critical_path_coverage_summary": {
            "critical_path_count": len(critical_path_records),
            "failing_final_paths": _string_list(critical_path_coverage.get("failing_final_paths")),
            "failing_fallback_paths": _string_list(critical_path_coverage.get("failing_fallback_paths")),
        },
        "fusion_summary": {
            "trusted_page_mapping_count": len(trusted_page_mappings),
            "provisional_mapping_count": len(provisional_mappings),
            "conflict_count": len(conflicting_groups),
            "unresolved_ambiguity_count": len(unresolved_ambiguities),
        },
        "delivery_readiness_breakdown": {
            "contract_version": "2026-05-26",
            "overall_status": audited_status,
            "final_gate_pass": final_delivery_ready,
            "fallback_gate_pass": fallback_gate_pass,
            "normal_mode_target": "90%+ critical coverage with 99%-100% confidence",
            "fallback_target": "85%+ bounded confidence without pretending to be final",
            "failing_final_gates": failing_final_gates,
            "failing_fallback_gates": failing_fallback_gates,
            "next_best_action": required_actions[0] if required_actions else "",
            "gates": gates,
        },
        "run_id": run_id,
        "skill_name": skill_name,
        "skill_version": skill_version,
        "stage_id": stage_id,
    }

    bounded_issue_pass_path = bundle_dir / "bounded_issue_pass.md"
    bounded_issue_pass_path.write_text(
        _render_bounded_issue_pass(
            safe_main_issues=safe_main_issues,
            demoted_main_issues=demoted_main_issues,
            audited_delivery_assessment=audited_delivery_assessment,
        ),
        encoding="utf-8",
    )

    unverified_path = bundle_dir / "unverified_issues.json"
    unverified_path.write_text(
        json.dumps(merged_unverified, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )

    supplement_request_path = bundle_dir / "supplement_request.md"
    supplement_request_path.write_text(
        _render_supplement_request(
            audited_delivery_assessment=audited_delivery_assessment,
            evidence_assessment=evidence,
            key_unverified_issues=key_unverified,
            demoted_main_issues=demoted_main_issues,
        ),
        encoding="utf-8",
    )

    audited_assessment_path = bundle_dir / "audited_delivery_assessment.json"
    audited_assessment_path.write_text(
        json.dumps(audited_delivery_assessment, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )

    benchmark_summary = _build_client_mode_benchmark_summary(
        output_dir=Path(output_dir),
        bundle_dir=bundle_dir,
        issues=issues,
        demoted_main_issues=demoted_main_issues,
        key_unverified_issues=key_unverified,
        evidence_assessment=evidence,
        audited_delivery_assessment=audited_delivery_assessment,
    )
    _write_client_mode_benchmark_summary(
        output_dir=Path(output_dir),
        summary=benchmark_summary,
    )
    audited_delivery_assessment["client_mode_metrics"] = benchmark_summary["metrics"]
    audited_delivery_assessment["benchmark_summary"] = benchmark_summary
    audited_assessment_path.write_text(
        json.dumps(audited_delivery_assessment, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )

    audit_manifest = {
        "generated_at": datetime.now(UTC).isoformat(),
        "delivery_status": audited_status,
        "qualified_issue_count": len(safe_main_issues),
        "demoted_main_issue_count": len(demoted_main_issues),
        "unverified_issue_count": len(merged_unverified),
        "files": [
            "bounded_issue_pass.md",
            "unverified_issues.json",
            "supplement_request.md",
            "audited_delivery_assessment.json",
            "../benchmark/client_mode_benchmark_summary.json",
            "../benchmark/client_mode_benchmark_summary.md",
        ],
    }
    (bundle_dir / "manifest.json").write_text(
        json.dumps(audit_manifest, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )

    return {
        "audited_delivery_assessment": audited_delivery_assessment,
        "client_mode_benchmark_summary": benchmark_summary,
        "delivery_audit_bundle": _artifact_payload(
            artifact_id="delivery_audit_bundle",
            output_type="delivery_audit_bundle",
            path=bundle_dir,
            format="directory",
            summary=(
                "Runtime-audited bounded fallback package."
                if audited_status != "final_delivery_ready"
                else "Runtime delivery audit bundle for a final-ready run."
            ),
            file_count=5,
            delivery_status=audited_status,
            qualified_issue_count=len(safe_main_issues),
            demoted_main_issue_count=len(demoted_main_issues),
        ),
    }


def build_issue_report(
    issues: list[dict[str, Any]],
    output_path: str | None = None,
    template: str = "uxeval",
    journey_map: Any = None,
    principles: Any = None,
    output_dir: str | None = None,
    run_id: str | None = None,
    skill_name: str | None = None,
    skill_version: str | None = None,
    stage_id: str | None = None,
) -> dict[str, Any]:
    """Build an Excel report from a list of issues.

    Args:
        issues: List of Issue objects serialized as dicts.
        output_path: Absolute path where the Excel file should be written.
            If None, generates a timestamped path in /tmp.
        template: Report template name ('uxeval', 'design-acceptance', 'competitor').
        journey_map: Journey map context included in html/evidence outputs.
        principles: Principles list included in html/evidence outputs.
        output_dir: Optional output directory used to derive stable artifact paths.

    Returns:
        Dict with ``issue_report`` / ``html_report`` / ``evidence_pack`` payloads.

    Raises:
        ExcelBuilderError: If template is unknown, path is invalid, or issues is empty.
    """
    # Validate template
    if template not in ["uxeval", "design-acceptance", "competitor"]:
        raise ExcelBuilderError(f"Unknown template: {template}")

    # Generate default output path if not provided
    if output_path is None:
        if output_dir is not None:
            output_path = str(Path(output_dir) / "issue_report.xlsx")
        else:
            timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
            output_path = f"/tmp/issue_report_{timestamp}.xlsx"

    # Validate output path
    output_file = Path(output_path)
    if not output_file.parent.exists():
        raise ExcelBuilderError(f"Output directory does not exist: {output_file.parent}")

    # Allow empty issues for some templates
    if not issues and template != "competitor":
        raise ExcelBuilderError("Issues list cannot be empty for this template")

    # Create workbook
    wb = Workbook()

    # Build template-specific sheets
    if template == "uxeval":
        _build_uxeval_template(wb, issues)
    elif template == "design-acceptance":
        _build_design_acceptance_template(wb, issues)
    elif template == "competitor":
        _build_competitor_template(wb, issues)

    # Save workbook
    wb.save(output_path)

    html_path = output_file.with_suffix(".html")
    _render_html_report(
        issues,
        target_path=html_path,
        journey_map=journey_map,
        principles=principles,
    )

    evidence_dir = output_file.parent / "evidence_pack"
    evidence_file_count = _write_evidence_pack(
        evidence_dir,
        issues=issues,
        journey_map=journey_map,
        principles=principles,
        excel_path=output_file,
        html_path=html_path,
        template=template,
    )

    return {
        "issue_report": _artifact_payload(
            artifact_id="issue_report",
            output_type="issue_report",
            path=output_file,
            format="xlsx",
            summary=f"{template} issue workbook with {len(issues)} issues.",
            sheet_count=len(wb.sheetnames),
            issue_count=len(issues),
        ),
        "html_report": _artifact_payload(
            artifact_id="html_report",
            output_type="html_report",
            path=html_path,
            format="html",
            summary=f"{template} HTML report with {len(issues)} issues.",
            issue_count=len(issues),
        ),
        "evidence_pack": _artifact_payload(
            artifact_id="evidence_pack",
            output_type="evidence_pack",
            path=evidence_dir,
            format="directory",
            summary="Evidence pack with manifest and structured source payloads.",
            file_count=evidence_file_count,
        ),
    }


__all__ = ["audit_delivery_readiness", "build_issue_report", "ExcelBuilderError"]
